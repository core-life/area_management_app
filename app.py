import secrets
import string
import os
import io
import pandas as pd
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, session, Response
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import or_, func, inspect # inspectをインポート
from sqlalchemy.exc import IntegrityError # IntegrityErrorをインポート
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import logging # loggingモジュールをインポート

# --- ロギングの設定 ---
# アプリケーションのログレベルを設定 (開発中はDEBUG、本番ではINFO/WARNINGなど)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
app_logger = logging.getLogger(__name__)

# --- Flaskアプリケーションの初期設定 ---
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your_super_secret_key_for_test')
# RenderでPostgreSQLを使う場合は、環境変数DATABASE_URLを設定します。
# SQLiteを使う場合は、下記の設定のままですが、データは永続化されません。
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///area_management.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# --- データベースモデルの定義 ---
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False) # ユーザー名はメールアドレスに固定
    name = db.Column(db.String(80), nullable=False) # 表示用の名前 (スペース除去済み)
    password_hash = db.Column(db.String(256), nullable=False) # パスワードをハッシュ化して保存
    affiliation = db.Column(db.String(100), nullable=True) # 所属
    is_admin = db.Column(db.Boolean, default=False) # 事務職員判定用
    is_first_login = db.Column(db.Boolean, default=True) # 初回ログインフラグ
    last_area_update = db.Column(db.DateTime, nullable=True) # 最終エリア更新日時を追加

    # パスワードハッシュ化のメソッド
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    # パスワード検証のメソッド
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def __repr__(self):
        return f'<User {self.email}>'

class Municipality(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    postal_code = db.Column(db.String(10), nullable=True)
    local_gov_code = db.Column(db.String(10), nullable=False, unique=True)
    prefecture = db.Column(db.String(20), nullable=False)
    city_town_village = db.Column(db.String(50), nullable=False)

    def __repr__(self):
        return '<Municipality %s %s>' % (self.prefecture, self.city_town_village)

class UserArea(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    municipality_id = db.Column(db.Integer, db.ForeignKey('municipality.id'), nullable=False)
    
    user = db.relationship('User', backref=db.backref('user_areas', lazy=True))
    municipality = db.relationship('Municipality', backref=db.backref('user_areas', lazy=True))

    def __repr__(self):
        return '<UserArea User:%s Muni:%s>' % (self.user_id, self.municipality_id)

class AreaChangeLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    municipality_id = db.Column(db.Integer, db.ForeignKey('municipality.id'), nullable=False)
    change_type = db.Column(db.String(20), nullable=False) # 'assigned' (対応可) or 'unassigned' (対応不可)
    change_date = db.Column(db.DateTime, nullable=False, default=datetime.utcnow) # 変更日時

    user = db.relationship('User', backref=db.backref('area_change_logs', lazy=True))
    municipality = db.relationship('Municipality', backref=db.backref('area_change_logs', lazy=True))

    def __repr__(self):
        return f'<AreaChangeLog User:{self.user_id} Muni:{self.municipality_id} Type:{self.change_type} Date:{self.change_date}>'


# --- データベースの初期化とデータ投入関数 ---
def init_db_and_data():
    with app.app_context():
        app_logger.info("データベースの初期化とデータ投入を開始します。")
        
        # データベースの状態をチェック
        inspector = inspect(db.engine)
        # 少なくともuserテーブルが存在しない場合、または全てのテーブルが存在しない場合、テーブルを再作成
        # これはRenderのEphemeral filesystem上のSQLiteでのみ機能し、データは永続化されません。
        # PostgreSQLを使う場合は、初回デプロイ時にのみdb.create_all()を実行するように調整が必要です。
        if not inspector.has_table("user"):
            app_logger.warning("データベーススキーマが不完全、または存在しません。テーブルを再作成します。")
            # 既存のテーブルをすべて削除してから作成（開発/デモ目的でのクリーンスタートに有効）
            db.drop_all() 
            db.create_all() 
            app_logger.info("データベーステーブルが作成されました。")
        else:
            app_logger.info("データベーステーブルは既に存在します。")

        # テスト管理者ユーザーの追加 (初回実行時のみ)
        admin_email = 'admin@clp-ytmm.com'
        if not User.query.filter_by(email=admin_email).first():
            admin_user = User(
                email=admin_email,
                name='管理者',
                affiliation='本社',
                is_admin=True,
                is_first_login=False,
                last_area_update=datetime.utcnow()
            )
            admin_user.set_password('admin_password')
            try:
                db.session.add(admin_user)
                db.session.commit()
                app_logger.info(f"管理者ユーザー {admin_email} を追加しました。")
            except IntegrityError:
                db.session.rollback()
                app_logger.warning(f"管理者ユーザー {admin_email} は既に存在します。")
            except Exception as e:
                db.session.rollback()
                app_logger.error(f"管理者ユーザーの追加中に予期せぬエラーが発生しました: {e}")
        else:
            app_logger.info(f"管理者ユーザー {admin_email} は既に存在します。")

        # 古いテスト営業職員ユーザーが存在する場合は削除する（モデル変更のため）
        old_test_sales_user = User.query.filter_by(name='test_user_sales').first()
        if old_test_sales_user:
            try:
                UserArea.query.filter_by(user_id=old_test_sales_user.id).delete()
                AreaChangeLog.query.filter_by(user_id=old_test_sales_user.id).delete()
                db.session.delete(old_test_sales_user)
                db.session.commit()
                app_logger.info("古い 'test_user_sales' ユーザーと関連データを削除しました。")
            except Exception as e:
                db.session.rollback()
                app_logger.error(f"古いテスト営業職員ユーザーの削除中にエラーが発生しました: {e}")
        db.session.commit() # 変更をコミット

        # 市区町村データの読み込みと投入 (初回実行時のみ)
        csv_file_path = 'municipalities.csv'
        
        if not Municipality.query.first() and os.path.exists(csv_file_path):
            app_logger.info(f"{csv_file_path}から市区町村データを投入します...")
            df = None
            read_csv_params = {'dtype': {'地方公共団体コード': str, '郵便番号': str}}
            try:
                df = pd.read_csv(csv_file_path, encoding='utf-8', **read_csv_params)
            except UnicodeDecodeError:
                app_logger.warning("utf-8での読み込みに失敗しました。cp932 (Shift-JIS) で再試行します。")
                try:
                    df = pd.read_csv(csv_file_path, encoding='cp932', **read_csv_params)
                except Exception as e:
                    app_logger.error(f"CSVファイルの読み込みエラー: cp932 (Shift-JIS) でも読み込みに失敗しました。ファイルが壊れているか、別のエンコーディングの可能性があります。エラー: {e}")
                    return
            except Exception as e:
                app_logger.error(f"CSVファイルの読み込み中に予期せぬエラーが発生しました: {e}")
                return

            if df is not None:
                expected_columns = ['郵便番号', '地方公共団体コード', '都道府県', '市区町村']
                if not all(col in df.columns for col in expected_columns):
                    app_logger.error(f"CSVファイルの列が期待と異なります。期待される列: {expected_columns}, 実際の列: {df.columns.tolist()}")
                    return

                for index, row in df.iterrows():
                    municipality = Municipality(
                        postal_code=row['郵便番号'],
                        local_gov_code=row['地方公共団体コード'],
                        prefecture=row['都道府県'],
                        city_town_village=row['市区町村']
                    )
                    db.session.add(municipality)
                try:
                    db.session.commit()
                    app_logger.info("市区町村データの投入が完了しました。")
                except IntegrityError:
                    db.session.rollback()
                    app_logger.error("市区町村データに重複する地方公共団体コードがあります。投入をスキップしました。")
                except Exception as e:
                    db.session.rollback()
                    app_logger.error(f"市区町村データのデータベースコミット中にエラーが発生しました: {e}")
            else:
                app_logger.warning("市区町村データの読み込みに失敗したため、データベースへの投入をスキップします。")
        elif not os.path.exists(csv_file_path):
            app_logger.warning(f"WARNING: {csv_file_path}が見つかりません。市区町村データが投入されません。")
        else:
            app_logger.info("市区町村データは既に投入されています。")

# アプリケーションのロード時にデータベース初期化を確実に実行
# GunicornがFlaskアプリをロードする際にこの部分が実行されます。
with app.app_context():
    init_db_and_data()

# --- ルート（URLと関数のマッピング）の定義 ---

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return render_template('login.html')

    if request.method == 'POST':
        email = request.form['email'].strip()
        password = request.form['password']
        user = User.query.filter_by(email=email).first()

        if user and user.check_password(password):
            session['user_id'] = user.id
            session['user_email'] = user.email
            session['user_name'] = user.name
            session['is_admin'] = user.is_admin
            app_logger.info(f"ユーザー {user.email} がログインしました。")
            
            if user.is_first_login:
                flash('初回ログインです。新しいパスワードを設定してください。', 'info')
                return redirect(url_for('reset_password'))
            elif user.is_admin:
                flash('ログインしました！', 'success')
                return redirect(url_for('admin_dashboard'))
            else:
                flash('ログインしました！', 'success')
                return redirect(url_for('sales_dashboard'))
        else:
            app_logger.warning(f"ログイン失敗: メールアドレス '{email}' またはパスワードが間違っています。")
            flash('メールアドレスまたはパスワードが間違っています。', 'danger')
            return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        email = request.form['email'].strip()
        affiliation = request.form['affiliation'].strip()
        name = request.form['name'].replace(' ', '').replace('　', '').strip() 

        if not email.endswith('@clp-ytmm.com'):
            flash('メールアドレスは "@clp-ytmm.com" ドメインである必要があります。', 'danger')
            app_logger.warning(f"ユーザー登録失敗: 不正なドメインのメールアドレス '{email}'。")
            return render_template('register.html', email=email, affiliation=affiliation, name=name)

        if User.query.filter_by(email=email).first():
            flash('このメールアドレスは既に登録されています。', 'danger')
            app_logger.warning(f"ユーザー登録失敗: メールアドレス '{email}' は既に登録済み。")
            return render_template('register.html', email=email, affiliation=affiliation, name=name)

        temporary_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for i in range(8))

        new_user = User(
            email=email,
            name=name,
            affiliation=affiliation,
            is_admin=False,
            is_first_login=True,
            last_area_update=None
        )
        new_user.set_password(temporary_password) 

        try:
            db.session.add(new_user)
            db.session.commit()
            app_logger.info(f"新規ユーザー {email} が登録されました。仮パスワード: {temporary_password} (非表示)") # Log the password, but don't show to user
            flash(f'ユーザー登録が完了しました。仮パスワードを <span class="font-semibold">{email}</span> 宛に送信しました。初回ログイン時にパスワードを変更してください。', 'success')
            return render_template('registration_success.html', email=email) # Removed temporary_password
        except IntegrityError:
            db.session.rollback()
            flash('このメールアドレスは既に登録されています。', 'danger')
            app_logger.error(f"ユーザー登録のコミット中にIntegrityErrorが発生: '{email}'")
            return render_template('register.html', email=email, affiliation=affiliation, name=name)
        except Exception as e:
            db.session.rollback()
            flash('ユーザー登録中に予期せぬエラーが発生しました。', 'danger')
            app_logger.error(f"ユーザー登録中に予期せぬエラーが発生: {e}", exc_info=True)
            return render_template('register.html', email=email, affiliation=affiliation, name=name)
    return render_template('register.html')

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    if 'user_id' not in session:
        flash('ログインしてください。', 'warning')
        return redirect(url_for('login'))

    user = User.query.get(session['user_id'])
    
    if not user:
        flash('ユーザー情報が見つかりませんでした。再度ログインしてください。', 'danger')
        app_logger.error(f"パスワードリセット失敗: セッションID {session['user_id']} のユーザーが見つかりません。")
        return redirect(url_for('login'))
    elif not user.is_first_login:
        flash('パスワードは既に設定済みです。再度パスワードをリセットするには、「パスワードを忘れた場合」をご利用ください。', 'info')
        app_logger.info(f"ユーザー {user.email} が初回ログインパスワードリセットを試みましたが、既に設定済みでした。")
        if user.is_admin:
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('sales_dashboard'))

    if request.method == 'POST':
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']

        if new_password != confirm_password:
            flash('新しいパスワードと確認用パスワードが一致しません。', 'danger')
            app_logger.warning(f"ユーザー {user.email} のパスワードリセット失敗: パスワード不一致。")
            return render_template('reset_password.html')

        try:
            user.set_password(new_password)
            user.is_first_login = False
            db.session.commit()
            app_logger.info(f"ユーザー {user.email} のパスワードが正常に更新されました。")
            flash('パスワードが正常に更新されました！', 'success')
            if user.is_admin:
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('sales_dashboard'))
        except Exception as e:
            db.session.rollback()
            flash('パスワード更新中にエラーが発生しました。', 'danger')
            app_logger.error(f"ユーザー {user.email} のパスワード更新中にエラーが発生: {e}", exc_info=True)
            return render_template('reset_password.html')
    return render_template('reset_password.html')

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form['email'].strip()

        if not email.endswith('@clp-ytmm.com'):
            flash('メールアドレスは "@clp-ytmm.com" ドメインである必要があります。', 'danger')
            app_logger.warning(f"パスワード忘れリクエスト失敗: 不正なドメインのメールアドレス '{email}'。")
            return render_template('forgot_password.html', email=email)

        user = User.query.filter_by(email=email).first()
        if user:
            temporary_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for i in range(8))
            try:
                user.set_password(temporary_password)
                user.is_first_login = True
                db.session.commit()
                app_logger.info(f"ユーザー {email} の仮パスワードが発行されました: {temporary_password} (非表示)") # Log the password, but don't show to user
                flash('パスワードリセットリクエストを受け付けました。新しい仮パスワードを <span class="font-semibold">{email}</span> 宛に送信しました。', 'success')
                return render_template('forgot_password_success.html', email=email) # Removed temporary_password
            except Exception as e:
                db.session.rollback()
                flash('パスワードリセット中にエラーが発生しました。', 'danger')
                app_logger.error(f"ユーザー {email} のパスワードリセット中にエラーが発生: {e}", exc_info=True)
                return render_template('forgot_password.html', email=email)
        else:
            flash('指定されたメールアドレスのユーザーは見つかりませんでした。', 'danger')
            app_logger.warning(f"パスワード忘れリクエスト失敗: メールアドレス '{email}' のユーザーが見つかりません。")
    return render_template('forgot_password.html')

@app.route('/logout')
def logout():
    user_email = session.get('user_email', '不明なユーザー')
    session.pop('user_id', None)
    session.pop('user_email', None)
    session.pop('user_name', None)
    session.pop('is_admin', None)
    flash('ログアウトしました。', 'info')
    app_logger.info(f"ユーザー {user_email} がログアウトしました。")
    return redirect(url_for('login'))

@app.route('/sales_dashboard', methods=['GET'])
def sales_dashboard():
    if 'user_id' not in session or session.get('is_admin'):
        flash('権限がありません。', 'danger')
        app_logger.warning(f"不正アクセス試行: 非ログインまたは管理者が営業職員ダッシュボードにアクセスしようとしました。User ID: {session.get('user_id')}")
        return redirect(url_for('login'))
    
    current_user_id = session['user_id']
    current_user = User.query.get(current_user_id)

    update_message = None
    if current_user and current_user.last_area_update:
        current_utc_date = datetime.utcnow()
        if current_user.last_area_update.year == current_utc_date.year and \
           current_user.last_area_update.month == current_utc_date.month:
            update_message = f"今月は既にエリアを更新しています。（最終更新: {current_user.last_area_update.strftime('%Y年%m月%d日 %H:%M')}）"
            flash(update_message, 'info')
            app_logger.info(f"ユーザー {current_user.email} は今月既にエリアを更新しています。")

    municipalities = Municipality.query.order_by(Municipality.local_gov_code).all()
    user_areas = UserArea.query.filter_by(user_id=current_user_id).all()
    user_selected_municipality_ids = {ua.municipality_id for ua in user_areas}

    prefecture_codes = db.session.query(
        Municipality.prefecture,
        func.min(Municipality.local_gov_code).label('min_local_gov_code')
    ).group_by(Municipality.prefecture).order_by('min_local_gov_code').all()
    
    prefectures = [p.prefecture for p in prefecture_codes]

    return render_template(
        'sales_dashboard.html',
        municipalities=municipalities,
        user_selected_municipality_ids=user_selected_municipality_ids,
        prefectures=prefectures,
        update_message=update_message
    )

@app.route('/save_sales_area', methods=['POST'])
def save_sales_area():
    if 'user_id' not in session or session.get('is_admin'):
        flash('権限がありません。', 'danger')
        app_logger.warning(f"不正アクセス試行: 非ログインまたは管理者が営業職員エリア保存を試みました。User ID: {session.get('user_id')}")
        return redirect(url_for('login'))

    current_user_id = session['user_id']
    current_user = User.query.get(current_user_id)

    if current_user.last_area_update:
        current_utc_date = datetime.utcnow()
        if current_user.last_area_update.year == current_utc_date.year and \
           current_user.last_area_update.month == current_utc_date.month:
            flash('今月は既にエリアを更新しています。エリア更新は月1回までです。', 'danger')
            app_logger.warning(f"ユーザー {current_user.email} が月次エリア更新制限に抵触しました。")
            return redirect(url_for('sales_dashboard'))

    selected_municipality_ids_str = request.form.getlist('selected_areas')
    selected_municipality_ids = {int(mid) for mid in selected_municipality_ids_str}

    current_user_areas = UserArea.query.filter_by(user_id=current_user_id).all()
    current_user_area_ids = {ua.municipality_id for ua in current_user_areas}

    areas_to_delete = current_user_area_ids - selected_municipality_ids
    areas_to_add = selected_municipality_ids - current_user_area_ids

    try:
        for muni_id in areas_to_delete:
            area_to_delete = UserArea.query.filter_by(user_id=current_user_id, municipality_id=muni_id).first()
            if area_to_delete:
                db.session.delete(area_to_delete)
                log_entry = AreaChangeLog(user_id=current_user_id, municipality_id=muni_id, change_type='unassigned')
                db.session.add(log_entry)
                app_logger.info(f"ユーザー {current_user.email} がエリア {muni_id} を担当不可に設定しました。")

        for muni_id in areas_to_add:
            new_user_area = UserArea(user_id=current_user_id, municipality_id=muni_id)
            db.session.add(new_user_area)
            log_entry = AreaChangeLog(user_id=current_user_id, municipality_id=muni_id, change_type='assigned')
            db.session.add(log_entry)
            app_logger.info(f"ユーザー {current_user.email} がエリア {muni_id} を担当可に設定しました。")

        if areas_to_delete or areas_to_add:
            current_user.last_area_update = datetime.utcnow()
            
        db.session.commit()
        flash('対応エリアが更新されました！', 'success')
        app_logger.info(f"ユーザー {current_user.email} の対応エリア更新が完了しました。")
    except Exception as e:
        db.session.rollback()
        flash('対応エリアの更新中にエラーが発生しました。', 'danger')
        app_logger.error(f"ユーザー {current_user.email} の対応エリア更新中にエラーが発生: {e}", exc_info=True)
    
    return redirect(url_for('sales_dashboard'))

@app.route('/sales_history')
def sales_history():
    if 'user_id' not in session or session.get('is_admin'):
        flash('権限がありません。', 'danger')
        app_logger.warning(f"不正アクセス試行: 非ログインまたは管理者が営業職員変更履歴にアクセスしようとしました。User ID: {session.get('user_id')}")
        return redirect(url_for('login'))

    current_user_id = session['user_id']
    one_year_ago = datetime.utcnow() - timedelta(days=365)

    try:
        history_logs = db.session.query(AreaChangeLog, Municipality, User).\
            join(Municipality, AreaChangeLog.municipality_id == Municipality.id).\
            join(User, AreaChangeLog.user_id == User.id).\
            filter(AreaChangeLog.user_id == current_user_id).\
            filter(AreaChangeLog.change_date >= one_year_ago).\
            order_by(AreaChangeLog.change_date.desc()).all()

        formatted_logs = []
        for log, muni, user in history_logs:
            formatted_logs.append({
                'change_date': log.change_date.strftime('%Y年%m月%d日 %H:%M'), 
                'prefecture': muni.prefecture,
                'city_town_village': muni.city_town_village,
                'change_type': '対応可' if log.change_type == 'assigned' else '対応不可'
            })
        app_logger.info(f"ユーザー {session.get('user_email')} の変更履歴表示に成功しました。")
    except Exception as e:
        app_logger.error(f"ユーザー {session.get('user_email')} の変更履歴取得中にエラーが発生: {e}", exc_info=True)
        flash('変更履歴の取得中にエラーが発生しました。', 'danger')
        formatted_logs = [] # エラー時は空のリストを返す

    return render_template('sales_history.html', history=formatted_logs)

@app.route('/admin_dashboard')
def admin_dashboard():
    if 'user_id' not in session or not session.get('is_admin'):
        flash('管理者権限が必要です。', 'danger')
        app_logger.warning(f"不正アクセス試行: 非管理者が管理者ダッシュボードにアクセスしようとしました。User ID: {session.get('user_id')}")
        return redirect(url_for('login'))
    
    search_prefecture = request.args.get('search_prefecture', '').strip()
    search_city_town_village = request.args.get('search_city_town_village', '').strip()
    search_affiliation = request.args.get('search_affiliation', '').strip()
    search_user_name = request.args.get('search_user_name', '').strip()

    try:
        municipalities_query = Municipality.query

        if search_prefecture:
            municipalities_query = municipalities_query.filter(
                Municipality.prefecture.ilike(f'%{search_prefecture}%')
            )
        if search_city_town_village:
            municipalities_query = municipalities_query.filter(
                Municipality.city_town_village.ilike(f'%{search_city_town_village}%')
            )

        all_municipalities = municipalities_query.order_by(
            Municipality.local_gov_code
        ).all()

        users_query = User.query.filter_by(is_admin=False)

        if search_user_name:
            users_query = users_query.filter(
                User.name.ilike(f'%{search_user_name}%')
            )
        if search_affiliation:
            users_query = users_query.filter(
                User.affiliation.ilike(f'%{search_affiliation}%')
            )
        
        all_users = users_query.order_by(User.name).all()

        municipality_user_map = {muni.id: set() for muni in all_municipalities}

        user_ids_to_display = [u.id for u in all_users]
        if user_ids_to_display:
            user_areas_for_displayed_municipalities = UserArea.query.filter(
                UserArea.user_id.in_(user_ids_to_display)
            ).all()
            for user_area in user_areas_for_displayed_municipalities:
                if user_area.municipality_id in municipality_user_map:
                    municipality_user_map[user_area.municipality_id].add(user_area.user_id)
        
        app_logger.info(f"管理者ダッシュボードの表示に成功しました。検索条件: 都道府県='{search_prefecture}', 市区町村='{search_city_town_village}', 担当者名='{search_user_name}', 所属='{search_affiliation}'")
    except Exception as e:
        app_logger.error(f"管理者ダッシュボードのデータ取得中にエラーが発生: {e}", exc_info=True)
        flash('データの取得中にエラーが発生しました。', 'danger')
        all_municipalities = []
        all_users = []
        municipality_user_map = {}

    return render_template(
        'admin_dashboard.html',
        all_municipalities=all_municipalities,
        all_users=all_users,
        municipality_user_map=municipality_user_map,
        search_prefecture=search_prefecture,
        search_city_town_village=search_city_town_village,
        search_user_name=search_user_name,
        search_affiliation=search_affiliation
    )

@app.route('/admin_users')
def admin_users():
    if 'user_id' not in session or not session.get('is_admin'):
        flash('管理者権限が必要です。', 'danger')
        app_logger.warning(f"不正アクセス試行: 非管理者がユーザー管理ページにアクセスしようとしました。User ID: {session.get('user_id')}")
        return redirect(url_for('login'))
    
    try:
        users = User.query.order_by(User.email).all()
        app_logger.info("ユーザーリストの表示に成功しました。")
    except Exception as e:
        app_logger.error(f"ユーザーリストの取得中にエラーが発生: {e}", exc_info=True)
        flash('ユーザーリストの取得中にエラーが発生しました。', 'danger')
        users = []

    return render_template('admin_users.html', users=users)

@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
def edit_user(user_id):
    if 'user_id' not in session or not session.get('is_admin'):
        flash('管理者権限が必要です。', 'danger')
        app_logger.warning(f"不正アクセス試行: 非管理者がユーザー編集ページにアクセスしようとしました。User ID: {session.get('user_id')}, 対象ユーザーID: {user_id}")
        return redirect(url_for('login'))
    
    user = User.query.get_or_404(user_id) 
    
    if user.id == session['user_id'] and user.is_admin:
        flash('ご自身の管理者権限は変更できません。', 'danger')
        app_logger.warning(f"ユーザー {session.get('user_email')} が自身の管理者権限変更を試みました。")
        return redirect(url_for('admin_users'))

    if request.method == 'POST':
        new_email = request.form['email'].strip() 
        new_name = request.form['name'].replace(' ', '').replace('　', '').strip() 
        new_affiliation = request.form['affiliation'].strip() 
        reset_password_flag = 'reset_password' in request.form 
        new_is_admin_status = 'is_admin' in request.form 

        if not new_email.endswith('@clp-ytmm.com'):
            flash('メールアドレスは "@clp-ytmm.com" ドメインである必要があります。', 'danger')
            app_logger.warning(f"ユーザー {user.email} の編集失敗: 不正なドメインのメールアドレス '{new_email}'。")
            return render_template('edit_user.html', user=user) 

        if new_email != user.email and User.query.filter_by(email=new_email).first():
            flash('このメールアドレスは既に他のユーザーに登録されています。', 'danger')
            app_logger.warning(f"ユーザー {user.email} の編集失敗: メールアドレス '{new_email}' は既に登録済み。")
            return render_template('edit_user.html', user=user) 

        user.email = new_email
        user.name = new_name
        user.affiliation = new_affiliation
        user.is_admin = new_is_admin_status 
        
        try:
            if reset_password_flag:
                temporary_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for i in range(8))
                user.set_password(temporary_password)
                user.is_first_login = True
                flash(f'ユーザー情報を更新し、パスワードをリセットしました。ユーザーには仮パスワードがメールで送信され、初回ログイン時に変更が必要です。', 'success')
                app_logger.info(f"ユーザー {user.email} のパスワードが管理者によってリセットされました。新しい仮パスワード: {temporary_password} (非表示)") # Log the password, but don't show to user
            else:
                flash('ユーザー情報が更新されました。', 'success')
                app_logger.info(f"ユーザー {user.email} の情報が管理者によって更新されました。")
            
            db.session.commit() 
            return redirect(url_for('admin_users')) 
        except IntegrityError:
            db.session.rollback()
            flash('メールアドレスの重複エラーが発生しました。', 'danger')
            app_logger.error(f"ユーザー {user.email} の編集コミット中にIntegrityErrorが発生: '{new_email}'")
            return render_template('edit_user.html', user=user)
        except Exception as e:
            db.session.rollback()
            flash('ユーザー情報の更新中にエラーが発生しました。', 'danger')
            app_logger.error(f"ユーザー {user.email} の情報更新中に予期せぬエラーが発生: {e}", exc_info=True)
            return render_template('edit_user.html', user=user)
    
    return render_template('edit_user.html', user=user)

@app.route('/delete_user/<int:user_id>', methods=['POST'])
def delete_user(user_id):
    if 'user_id' not in session or not session.get('is_admin'):
        flash('管理者権限が必要です。', 'danger')
        app_logger.warning(f"不正アクセス試行: 非管理者がユーザー削除を試みました。User ID: {session.get('user_id')}, 対象ユーザーID: {user_id}")
        return redirect(url_for('login'))
    
    user = User.query.get_or_404(user_id) 
    if user.id == session['user_id']:
        flash('ご自身のアカウントは削除できません。', 'danger')
        app_logger.warning(f"ユーザー {session.get('user_email')} が自身の削除を試みました。")
        return redirect(url_for('admin_users'))
    if user.is_admin: 
        flash('管理者ユーザーは削除できません。', 'danger')
        app_logger.warning(f"管理者 ({user.email}) の削除が阻止されました。")
        return redirect(url_for('admin_users'))
    
    try:
        UserArea.query.filter_by(user_id=user.id).delete()
        AreaChangeLog.query.filter_by(user_id=user.id).delete()
        db.session.delete(user) 
        db.session.commit() 
        flash(f'ユーザー "{user.name}" と関連データが削除されました。', 'success')
        app_logger.info(f"ユーザー {user.email} と関連データが削除されました。")
    except Exception as e:
        db.session.rollback()
        flash(f'ユーザー "{user.name}" の削除中にエラーが発生しました。', 'danger')
        app_logger.error(f"ユーザー {user.email} の削除中にエラーが発生: {e}", exc_info=True)
    return redirect(url_for('admin_users')) 

@app.route('/download_excel', defaults={'months': 12})
@app.route('/download_excel/<int:months>')
def download_excel(months):
    if 'user_id' not in session or not session.get('is_admin'):
        flash('管理者権限が必要です。', 'danger')
        app_logger.warning(f"不正アクセス試行: 非管理者がExcelダウンロードを試みました。User ID: {session.get('user_id')}")
        return redirect(url_for('login'))
    
    months_to_export = months
    if not (1 <= months_to_export <= 12):
        months_to_export = 12 

    current_date = datetime.utcnow()
    
    start_year = current_date.year
    start_month = current_date.month - (months_to_export - 1)
    while start_month <= 0:
        start_month += 12
        start_year -= 1
    start_date_of_history = datetime(start_year, start_month, 1, 0, 0, 0, 0)

    try:
        all_municipalities = Municipality.query.order_by(
            Municipality.local_gov_code
        ).all()

        all_users = User.query.filter_by(is_admin=False).order_by(User.name).all()

        output = io.BytesIO()
        wb = Workbook()
        
        ws_main = wb.active
        ws_main.title = '対応エリア一覧'

        no_border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')

        row1_cells = ['郵便番号', '地方公共団体コード', '住所①', '住所②']
        user_affiliations = [user.affiliation if user.affiliation else '' for user in all_users]
        row1_cells.extend(user_affiliations)
        ws_main.append(row1_cells)

        row2_cells = ['', '', '', '']
        user_names = [user.name for user in all_users]
        row2_cells.extend(user_names)
        ws_main.append(row2_cells)

        for row_idx in range(1, 3):
            for col_idx in range(1, len(row1_cells) + 1):
                cell = ws_main.cell(row=row_idx, column=col_idx)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = no_border

        ws_main.merge_cells('A1:A2')
        ws_main.merge_cells('B1:B2')
        ws_main.merge_cells('C1:C2')
        ws_main.merge_cells('D1:D2')

        for municipality in all_municipalities:
            row_data = [
                municipality.postal_code if municipality.postal_code else '',
                municipality.local_gov_code,
                municipality.prefecture,
                municipality.city_town_village
            ]
            for user in all_users:
                is_assigned = db.session.query(UserArea).filter_by(
                    user_id=user.id, municipality_id=municipality.id
                ).first() is not None
                row_data.append('〇' if is_assigned else '')
            ws_main.append(row_data)

        ws_main.column_dimensions['D'].width = 20.75

        for row_idx in range(3, ws_main.max_row + 1):
            for col_idx in range(1, ws_main.max_column + 1):
                cell = ws_main.cell(row=row_idx, column=col_idx)
                cell.border = no_border

        ws_history = wb.create_sheet('エリア変更履歴_月次')
        
        history_logs = db.session.query(AreaChangeLog, Municipality, User).\
            join(Municipality, AreaChangeLog.municipality_id == Municipality.id).\
            join(User, AreaChangeLog.user_id == User.id).\
            filter(AreaChangeLog.change_date >= start_date_of_history).\
            order_by(AreaChangeLog.change_date.asc()).all() 

        monthly_changes = {} 
        for log, muni, user in history_logs:
            change_month = log.change_date.strftime('%Y年%m月')
            user_name = user.name
            area_name = f"{muni.prefecture}{muni.city_town_village}"
            change_type = log.change_type

            if change_month not in monthly_changes:
                monthly_changes[change_month] = {}
            if user_name not in monthly_changes[change_month]:
                monthly_changes[change_month][user_name] = {'assigned': [], 'unassigned': []}
            
            monthly_changes[change_month][user_name][change_type].append(area_name)
        
        history_data_summary = []
        for month in sorted(monthly_changes.keys()):
            for user_name in sorted(monthly_changes[month].keys()):
                assigned_areas = "、".join(monthly_changes[month][user_name]['assigned'])
                unassigned_areas = "、".join(monthly_changes[month][user_name]['unassigned'])
                
                history_data_summary.append({
                    '対象月': month,
                    '営業職員名': user_name,
                    '対応可能エリア追加': assigned_areas,
                    '対応可能エリア削除': unassigned_areas
                })
        
        df_history = pd.DataFrame(history_data_summary)
        
        if not df_history.empty: # データがある場合のみシートを追加
            for r_idx, row in enumerate(dataframe_to_rows(df_history, index=False, header=True), 1):
                ws_history.append(row)
            
            for row in ws_history.iter_rows():
                for cell in row:
                    cell.border = no_border
        else: # 履歴がない場合でもシートを作成し、メッセージを記載
            ws_history.append(['選択された期間のエリア変更履歴はありません。'])
            if ws_history['A1']:
                ws_history['A1'].border = no_border

        wb.save(output)
        output.seek(0) 

        filename = f"area_list_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        app_logger.info(f"Excelファイル '{filename}' のダウンロードが要求されました。")
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
    except Exception as e:
        app_logger.error(f"Excelファイルの生成またはダウンロード中にエラーが発生: {e}", exc_info=True)
        flash('Excelファイルの生成中にエラーが発生しました。', 'danger')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin_upload_municipalities', methods=['GET', 'POST'])
def admin_upload_municipalities():
    if 'user_id' not in session or not session.get('is_admin'):
        flash('管理者権限が必要です。', 'danger')
        app_logger.warning(f"不正アクセス試行: 非管理者が市区町村データ一括更新ページにアクセスしようとしました。User ID: {session.get('user_id')}")
        return redirect(url_for('login'))

    additions = []
    updates = []
    deletions = []
    processing_error = None

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('ファイルが選択されていません。', 'danger')
            app_logger.warning("市区町村データアップロード失敗: ファイルが選択されていません。")
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            flash('ファイルが選択されていません。', 'danger')
            app_logger.warning("市区町村データアップロード失敗: ファイル名が空です。")
            return redirect(request.url)
        
        if file and file.filename.endswith('.csv'):
            try:
                file_content = io.BytesIO(file.read())
                df = None
                read_csv_params = {'dtype': {'地方公共団体コード': str, '郵便番号': str}} 
                
                try:
                    df = pd.read_csv(file_content, encoding='utf-8', **read_csv_params)
                except UnicodeDecodeError:
                    file_content.seek(0)
                    df = pd.read_csv(file_content, encoding='cp932', **read_csv_params)

                expected_columns = ['郵便番号', '地方公共団体コード', '都道府県', '市区町村'] 
                if not all(col in df.columns for col in expected_columns):
                    processing_error = f"CSVファイルの列が期待と異なります。期待される列: {expected_columns}, 実際の列: {df.columns.tolist()}"
                    app_logger.error(f"市区町村データCSVの列エラー: {processing_error}")
                
                if not processing_error and df['地方公共団体コード'].duplicated().any():
                    duplicate_codes = df[df['地方公共団体コード'].duplicated()]['地方公共団体コード'].tolist()
                    processing_error = f"CSVファイル内に重複する地方公共団体コードがあります: {', '.join(duplicate_codes)}"
                    app_logger.error(f"市区町村データCSVの重複コードエラー: {processing_error}")

                if not processing_error:
                    existing_municipalities = {
                        m.local_gov_code: m for m in Municipality.query.all()
                    }
                    existing_codes = set(existing_municipalities.keys())
                    new_codes = set()

                    for index, row in df.iterrows():
                        local_gov_code = str(row.get('地方公共団体コード', '')).strip()
                        if not local_gov_code: # 地方公共団体コードが空の場合はスキップ
                            app_logger.warning(f"CSVの行 {index+2} に地方公共団体コードがありません。この行はスキップされました。")
                            continue
                        new_codes.add(local_gov_code)

                        postal_code = str(row.get('郵便番号', '')).strip() if pd.notna(row.get('郵便番号')) else ''
                        prefecture = str(row.get('都道府県', '')).strip() if pd.notna(row.get('都道府県')) else ''
                        city_town_village = str(row.get('市区町村', '')).strip() if pd.notna(row.get('市区町村')) else ''

                        if local_gov_code in existing_codes:
                            existing_muni = existing_municipalities[local_gov_code]
                            if (existing_muni.postal_code != postal_code or
                                existing_muni.prefecture != prefecture or
                                existing_muni.city_town_village != city_town_village):
                                updates.append({
                                    'local_gov_code': local_gov_code,
                                    'old_postal_code': existing_muni.postal_code,
                                    'new_postal_code': postal_code,
                                    'old_prefecture': existing_muni.prefecture,
                                    'new_prefecture': prefecture,
                                    'old_city_town_village': existing_muni.city_town_village,
                                    'new_city_town_village': city_town_village
                                })
                        else:
                            additions.append({
                                'postal_code': postal_code,
                                'local_gov_code': local_gov_code,
                                'prefecture': prefecture,
                                'city_town_village': city_town_village
                            })
                    
                    deletions_codes = existing_codes - new_codes
                    for code in deletions_codes:
                        muni = existing_municipalities[code]
                        deletions.append({
                            'postal_code': muni.postal_code,
                            'local_gov_code': muni.local_gov_code,
                            'prefecture': muni.prefecture,
                            'city_town_village': muni.city_town_village
                        })
                    
                    session['pending_additions'] = additions
                    session['pending_updates'] = updates
                    session['pending_deletions'] = deletions

                    if not additions and not updates and not deletions:
                        flash('CSVファイルと既存のデータに差異はありませんでした。', 'info')
                        app_logger.info("市区町村データ更新プレビュー: 差異なし。")
                        return redirect(url_for('admin_dashboard'))
                    else:
                        flash('CSVデータを読み込みました。以下の変更が適用されます。内容を確認し「確定して実行」してください。', 'info')
                        app_logger.info(f"市区町村データ更新プレビュー表示: 追加 {len(additions)}件, 更新 {len(updates)}件, 削除 {len(deletions)}件。")

            except pd.errors.EmptyDataError:
                processing_error = 'CSVファイルが空です。'
                app_logger.error("市区町村データアップロードエラー: CSVファイルが空です。")
            except pd.errors.ParserError:
                processing_error = 'CSVファイルの解析に失敗しました。形式を確認してください。'
                app_logger.error(f"市区町村データアップロードエラー: CSVファイルの解析失敗。{e}", exc_info=True)
            except UnicodeDecodeError:
                processing_error = 'CSVファイルのエンコーディングがUTF-8またはShift-JISではありません。'
                app_logger.error("市区町村データアップロードエラー: 不明なエンコーディング。")
            except Exception as e:
                processing_error = f'CSV処理中にエラーが発生しました: {e}'
                app_logger.error(f"市区町村データアップロード中に予期せぬエラーが発生: {e}", exc_info=True)
            
            if processing_error:
                flash(f'CSV処理エラー: {processing_error}', 'danger')
                session.pop('pending_additions', None)
                session.pop('pending_updates', None)
                session.pop('pending_deletions', None)

        else:
            flash('CSVファイルを選択してください。', 'danger')
            app_logger.warning("市区町村データアップロード失敗: CSVファイル以外が選択されました。")

    return render_template(
        'admin_upload_municipalities.html',
        additions=session.get('pending_additions', []),
        updates=session.get('pending_updates', []),
        deletions=session.get('pending_deletions', [])
    )

@app.route('/admin_execute_municipality_update', methods=['POST'])
def admin_execute_municipality_update():
    if 'user_id' not in session or not session.get('is_admin'):
        flash('管理者権限が必要です。', 'danger')
        app_logger.warning(f"不正アクセス試行: 非管理者が市区町村データ更新実行を試みました。User ID: {session.get('user_id')}")
        return redirect(url_for('login'))

    additions = session.pop('pending_additions', [])
    updates = session.pop('pending_updates', [])
    deletions = session.pop('pending_deletions', [])

    if not additions and not updates and not deletions:
        flash('適用する変更がありません。', 'warning')
        app_logger.info("市区町村データ更新実行: 適用する変更がありませんでした。")
        return redirect(url_for('admin_dashboard'))

    try:
        for muni_data in deletions:
            muni_to_delete = Municipality.query.filter_by(local_gov_code=muni_data['local_gov_code']).first()
            if muni_to_delete:
                UserArea.query.filter_by(municipality_id=muni_to_delete.id).delete()
                AreaChangeLog.query.filter_by(municipality_id=muni_to_delete.id).delete()
                db.session.delete(muni_to_delete)
                app_logger.info(f"市区町村 '{muni_data['local_gov_code']}' と関連データが削除されました。")
        
        for item_data in additions:
            new_municipality = Municipality(
                postal_code=item_data['postal_code'],
                local_gov_code=item_data['local_gov_code'],
                prefecture=item_data['prefecture'],
                city_town_village=item_data['city_town_village']
            )
            db.session.add(new_municipality)
            app_logger.info(f"市区町村 '{item_data['local_gov_code']}' が追加されました。")
        
        for muni_data in updates:
            muni_to_update = Municipality.query.filter_by(local_gov_code=muni_data['local_gov_code']).first()
            if muni_to_update:
                muni_to_update.postal_code = muni_data['new_postal_code']
                muni_to_update.prefecture = muni_data['new_prefecture']
                muni_to_update.city_town_village = muni_data['new_city_town_village']
                app_logger.info(f"市区町村 '{muni_data['local_gov_code']}' が更新されました。")
        
        db.session.commit()
        flash(f'市区町村データが正常に更新されました！ (追加: {len(additions)}件, 更新: {len(updates)}件, 削除: {len(deletions)}件)', 'success')
        app_logger.info(f"市区町村データの一括更新が正常に完了しました。追加: {len(additions)}件, 更新: {len(updates)}件, 削除: {len(deletions)}件。")

    except IntegrityError as e:
        db.session.rollback()
        flash(f'市区町村データの更新中に整合性エラーが発生しました。重複する地方公共団体コードがないか確認してください。', 'danger')
        app_logger.error(f"市区町村データの更新中にIntegrityErrorが発生: {e}", exc_info=True)
        session['pending_additions'] = additions
        session['pending_updates'] = updates
        session['pending_deletions'] = deletions
    except Exception as e:
        db.session.rollback()
        flash(f'市区町村データの更新中にエラーが発生しました: {e}', 'danger')
        app_logger.error(f"市区町村データの更新中に予期せぬエラーが発生: {e}", exc_info=True)
        session['pending_additions'] = additions
        session['pending_updates'] = updates
        session['pending_deletions'] = deletions
    
    return redirect(url_for('admin_dashboard'))

@app.route('/')
def index():
    return redirect(url_for('login'))

if __name__ == '__main__':
    # ローカル開発環境でのみ、init_db_and_data()を明示的に呼び出す（自動実行はしない）
    # Renderでは、app.app_context()ブロックで既に呼び出されます。
    app.run(debug=True)
